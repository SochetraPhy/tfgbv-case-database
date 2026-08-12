#!/usr/bin/env python3
"""
Import the legacy TFGBV workbook (Template 1.xlsm) into Directus.

Dry run by default — reads the workbook, resolves every value, and reports exactly what
it *would* create, including anything it could not map. Nothing is written until you
pass --commit.

    python3 scripts/import_legacy.py "/path/to/workbook.xlsm"            # dry run
    python3 scripts/import_legacy.py "/path/to/workbook.xlsm" --commit   # write

Run this yourself against the populated workbook. It contains survivor and perpetrator
PII: keep it on an encrypted volume, outside this repo, and do not paste its contents
anywhere. The script writes PII only into survivor_pii / perpetrator_pii, which only
Super Admin can read.

Requires: pip3 install openpyxl

Notes on what it does
---------------------
* Multi-value cells ("Facebook, Instagram, Tik Tok") are split and resolved into the
  many-to-many junctions.
* Labels are matched case- and punctuation-insensitively, with an alias table for the
  drift between the workbook and the dashboard sheet ("Human Right Defender").
* Anything unmatched is REPORTED, never silently dropped or invented.
* One survivor and one perpetrator row per case. The workbook has no identity key that
  would let rows be safely merged, and wrongly merging two survivors is far worse than
  duplicating one.
* Re-running skips cases whose case_code already exists, so a partial import can be
  resumed.
"""
import argparse
import datetime as dt
import re
import sys
from collections import defaultdict

sys.path.insert(0, __file__.rsplit("/", 1)[0])
import seed_schema as s  # noqa: E402

MASTER_SHEET = "ODC_TFGBV Case Masterfile"
SUPPORT_SHEET = "Technical Support Case"

# Workbook label -> vocabulary key, for values that do not match a label directly.
ALIASES = {
    "identities": {
        "human right defender": "human_rights_defender",
        "human rights defenders": "human_rights_defender",
        "political activist/hrd": "political_activist",
        "lgbtqia+ communities": "lgbtqia_community",
        "people living with disabilities": "people_with_disabilities",
        "people with disailities": "people_with_disabilities",
        "n/a": None,          # explicitly "no value", not an error
        "na": None,
    },
    "genders": {
        "lgbtqia+ communities": "lgbtqia_community",
        "gender diverse": "gender_diverse_group",
        "n/a": None,
    },
    "platforms": {"tik tok": "tiktok", "tiktok": "tiktok", "n/a": None},
    "harassment_types": {"technical support": None},   # an intervention, not harassment
    "interventions": {"n/a": "not_applicable"},
    "case_statuses": {"under investigation": "investigating"},
    "perpetrator_relations": {"n/a": "not_applicable"},
    "case_categories": {},
    "severity_levels": {},
    "reporters": {},
}


def norm(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("’", "'").replace("–", "-")
    return re.sub(r"\s+", " ", text)


class Vocab:
    """Label -> key resolution for one vocabulary, with unmapped-value tracking."""

    def __init__(self, collection):
        self.collection = collection
        self.by_label = {}
        for row in s.api("GET", f"/items/{collection}?limit=-1&fields=id,label"):
            self.by_label[norm(row["label"])] = row["id"]
            self.by_label[norm(row["id"])] = row["id"]
        self.aliases = {norm(k): v for k, v in ALIASES.get(collection, {}).items()}
        self.unmapped = defaultdict(int)

    def resolve(self, value):
        key = norm(value)
        if not key:
            return None
        if key in self.aliases:
            return self.aliases[key]
        if key in self.by_label:
            return self.by_label[key]
        self.unmapped[str(value).strip()] += 1
        return None

    def resolve_many(self, cell):
        """Extract every known label from a multi-value cell.

        Deliberately not a split on separators. Real labels contain the separators you
        would naively split on — "Manager/Supervisor", "Artist/Celebrity/Public Figure",
        "N/A", and "Sextortion (blackmail,...)" with a comma inside it. Instead, match
        known labels longest-first at word boundaries and remove each as it is found;
        whatever text is left over is genuinely unrecognised and gets reported.
        """
        if cell is None:
            return []
        text = norm(cell)
        if not text:
            return []

        candidates = sorted(
            list(self.aliases.items()) + list(self.by_label.items()),
            key=lambda kv: -len(kv[0]))

        out = []
        for label, key in candidates:
            if not label:
                continue
            pattern = r"(?<!\w)" + re.escape(label) + r"(?!\w)"
            if re.search(pattern, text):
                text = re.sub(pattern, " ", text)
                if key and key not in out:
                    out.append(key)

        leftover = re.sub(r"[,;/&]|\band\b", " ", text)
        leftover = re.sub(r"\s+", " ", leftover).strip()
        if leftover and leftover not in {"-", "n a", "none"}:
            self.unmapped[leftover] += 1
        return out


def as_date(value):
    if value is None:
        return None
    if isinstance(value, (dt.datetime, dt.date)):
        return value.strftime("%Y-%m-%d")
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%b-%Y"):
        try:
            return dt.datetime.strptime(str(value).strip(), fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def m2m(field, keys):
    return {"create": [{f"{field}_id": {"id": k}} for k in keys],
            "update": [], "delete": []}


def rows_of(ws):
    header = [str(c) if c is not None else "" for c in
              next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(c is None for c in row):
            continue
        yield dict(zip(header, row))


def pick(row, *candidates):
    """Column headers vary slightly between the two sheets and across file versions."""
    for key in row:
        flat = norm(key)
        for candidate in candidates:
            if candidate in flat:
                return row[key]
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook")
    ap.add_argument("--commit", action="store_true",
                    help="actually write to Directus (default is a dry run)")
    ap.add_argument("--organization", help="organization name to attach cases to")
    args = ap.parse_args()

    try:
        import openpyxl
    except ImportError:
        raise SystemExit("openpyxl is required:  pip3 install openpyxl")

    s.login()
    V = {name: Vocab(name) for name in
         ("case_categories", "genders", "identities", "platforms", "harassment_types",
          "case_statuses", "interventions", "perpetrator_relations", "severity_levels",
          "reporters")}

    org_id = None
    orgs = s.api("GET", "/items/organizations?limit=-1&fields=id,name")
    if args.organization:
        match = [o for o in orgs if norm(o["name"]) == norm(args.organization)]
        if not match:
            raise SystemExit(f"no organization named {args.organization!r}")
        org_id = match[0]["id"]
    elif len(orgs) == 1:
        org_id = orgs[0]["id"]

    existing = {c["case_code"] for c in
                s.api("GET", "/items/cases?limit=-1&fields=case_code")}

    wb = openpyxl.load_workbook(args.workbook, data_only=True)
    planned, skipped, with_pii = [], 0, 0

    for sheet_name, case_type in ((MASTER_SHEET, "tfgbv"),
                                  (SUPPORT_SHEET, "technical_support")):
        if sheet_name not in wb.sheetnames:
            print(f"  (no sheet {sheet_name!r} — skipping)")
            continue
        for row in rows_of(wb[sheet_name]):
            code = pick(row, "case id")
            if not code:
                continue
            code = str(code).strip()
            if code in existing:
                skipped += 1
                continue

            survivor_name = pick(row, "victim's name", "survivor's name", "requester")
            known = pick(row, "known perpetrator")
            suspected = pick(row, "suspected perpetrator")

            case = {
                "case_code": code,
                "case_type": case_type,
                "record_status": "submitted",
                "date_reported": as_date(pick(row, "date reported")),
                "case_category": V["case_categories"].resolve(pick(row, "case category")),
                "case_status": V["case_statuses"].resolve(pick(row, "case status")),
                "severity": V["severity_levels"].resolve(pick(row, "severity")),
                "perpetrator_relation": V["perpetrator_relations"].resolve(
                    pick(row, "relationship to victim", "perpetrator  relationship")),
                "caption_main_content": pick(row, "caption written"),
                "incident_summary": pick(row, "summary of the incident"),
                "impact_on_survivor": pick(row, "impact on the victim"),
                "organization": org_id,
                "platforms": V["platforms"].resolve_many(pick(row, "platform")),
                "harassment_types": V["harassment_types"].resolve_many(
                    pick(row, "harrassment type", "harassment type")),
                "interventions": V["interventions"].resolve_many(
                    pick(row, "actions taken", "type of support")),
            }
            links = pick(row, "link to online")
            if links:
                case["evidence_links"] = [{"url": u.strip()} for u in
                                          re.split(r"[\s,;]+", str(links)) if u.strip()]

            survivor = {
                "gender": V["genders"].resolve(pick(row, "gender")),
                "identities": V["identities"].resolve_many(pick(row, "other identities")),
            }
            perpetrator = None
            if known or suspected:
                perpetrator = {
                    "identification_status": "known" if known else "suspected",
                    "known_information": known,
                    "suspected_information": suspected,
                }
            if survivor_name or known or suspected:
                with_pii += 1

            planned.append((code, case, survivor, survivor_name, perpetrator))

    # ------------------------------------------------------------------ reporting
    print(f"\nWorkbook: {args.workbook}")
    print(f"  cases to import : {len(planned)}")
    print(f"  already present : {skipped} (skipped)")
    print(f"  rows carrying PII: {with_pii} -> survivor_pii / perpetrator_pii")
    print(f"  organization    : {args.organization or '(single existing org)' if org_id else 'NONE — set with --organization'}")

    problems = False
    for name, vocab in V.items():
        if vocab.unmapped:
            problems = True
            print(f"\n  UNMAPPED in {name}:")
            for value, count in sorted(vocab.unmapped.items(), key=lambda kv: -kv[1]):
                print(f"     {count:4}x  {value!r}")

    missing_date = [c for c, case, *_ in planned if not case["date_reported"]]
    if missing_date:
        problems = True
        print(f"\n  MISSING date_reported ({len(missing_date)}): {missing_date[:10]}")
    if not org_id:
        problems = True
        print("\n  No organization resolved — submitted cases require one.")

    if problems:
        print("\nResolve the above first: add the value to the right vocabulary, or add"
              "\nan alias in ALIASES at the top of this script. Nothing is invented.")

    if not args.commit:
        print("\nDry run — nothing written. Re-run with --commit when the report is clean.")
        return
    if problems:
        raise SystemExit("\nRefusing to import while values are unresolved.")

    # --------------------------------------------------------------------- write
    for code, case, survivor, survivor_name, perpetrator in planned:
        # Track what this row created so a failure part-way does not strand a survivor
        # or perpetrator record with no case attached to it.
        created = []
        try:
            sv = s.api("POST", "/items/survivors", {
                "gender": survivor["gender"],
                "identities": m2m("identities", survivor["identities"])})
            created.append(("survivors", sv["id"]))
            if survivor_name:
                s.api("POST", "/items/survivor_pii",
                      {"survivor": sv["id"], "full_name": str(survivor_name).strip()})
            case["survivor"] = sv["id"]

            if perpetrator:
                pr = s.api("POST", "/items/perpetrators", {
                    "identification_status": perpetrator["identification_status"]})
                created.append(("perpetrators", pr["id"]))
                s.api("POST", "/items/perpetrator_pii", {
                    "perpetrator": pr["id"],
                    "known_information": perpetrator["known_information"],
                    "suspected_information": perpetrator["suspected_information"]})
                case["perpetrator"] = pr["id"]

            for field in ("platforms", "harassment_types", "interventions"):
                case[field] = m2m(field, case[field])
            s.api("POST", "/items/cases", case)
            print(f"  imported {code}")
        except SystemExit as e:
            for collection, item_id in reversed(created):
                s.api("DELETE", f"/items/{collection}/{item_id}", quiet404=True)
            raise SystemExit(f"\n{code} failed, its partial rows were rolled back:\n{e}")

    print(f"\nDone — {len(planned)} cases imported.")


if __name__ == "__main__":
    main()
